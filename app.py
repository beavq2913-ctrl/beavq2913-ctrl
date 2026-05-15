import re
import io
import zipfile
import unicodedata
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── REGLAS ──────────────────────────────────────────────────────────────────

PCT_GENERAL = 0.13
PATOLOGO_CERO = {
    "VANESA MIKOLAITIS","NURIA GALLEGO TREJO","SILVIA VIVIANA HERRERA",
    "VALERIA ALBERTON","ELAINE DE FREITAS","EXTERNO PATOLOGO",
}

# Patólogos inscriptos en IVA — para el resto el IVA es siempre 0
PATOLOGO_CON_IVA = {
    "FEDERICO BASILI",
    "ALEJANDRA AVAGNINA",
    "GISELLE ROMERO CAIMI",
    "CLAUDIO LEWIN",
    "VALERIA ALBERTON",
    "ANTONIO SACCO",
}
MONTO_FIJO_ALBERTON_DEFAULT = 85129.46  # valor base — se sobreescribe desde la UI
PATOLOGO_26 = {"CLAUDIO LEWIN"}
CODIGOS_94_PCT = {"150104","150115","150888","99999"}
CODIGOS_BARRIENTOS_FIJO = {"50104","150104","150115","150888","99999"}
MONTO_FIJO_BARRIENTOS = 25000.0
PATOLOGO_PROSTATA_ESPECIAL = {
    "DAIRA NAIMAN",
    "GABRIELA INES CREVENA",   # cubre variante INÉS
    "GABRIELA INƁ CREVENA",    # cubre variante con carácter IPA ɓ (ord=595)
    "NILDA GONZALEZ ROIBON",
}
COD_PROSTATA = "150103"
PRECIO_FIJO_BSI = 100000.0

def norm(val):
    if pd.isna(val) or str(val).strip() == "":
        return ""
    s = str(val).replace("\xa0", " ").strip().upper()
    s = " ".join(s.split())
    # Normalizar caracteres unicode (ɓ→B, É→E, etc.) para matching robusto
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s


def es_prostata(s):
    return bool(re.search(r"PR.{0,3}T.{0,2}TA", str(s).replace("\xa0", " ").upper()))

def calcular_importe(patologo, firma, subtotal, cod_fact, organo_raw, derivante, cantidad=1):
    """Retorna (importe, regla, es_presencia).
    es_presencia=True cuando el importe va a la columna Presencias en lugar de Importe Firma."""
    if not patologo:
        return 0.0, "Sin patólogo", False
    cod = str(cod_fact).strip()
    der = norm(derivante)

    if patologo == "GUSTAVO BARRIENTOS":
        if cod in CODIGOS_BARRIENTOS_FIJO:
            importe = MONTO_FIJO_BARRIENTOS * cantidad
            return importe, f"Barrientos {cod} → $25.000 × {cantidad} = ${importe:,.0f}", True

    if cod in CODIGOS_94_PCT and patologo != "GUSTAVO BARRIENTOS":
        if firma == "segunda":
            return 0.0, f"Código {cod} → 2da firma ignorada", False
        return subtotal * 0.94, f"Código {cod} → 94% subtotal", True

    if patologo == "VALERIA ALBERTON":
        return 0.0, "Valeria Alberton → monto fijo por ingreso", False

    if patologo in PATOLOGO_CERO:
        return 0.0, f"{patologo} → 0%", False

    if patologo in PATOLOGO_26:
        return subtotal * 0.26, f"{patologo} → 26%", False

    if patologo in PATOLOGO_PROSTATA_ESPECIAL:
        if cod == COD_PROSTATA and es_prostata(organo_raw):
            if der == "BSI":
                base = PRECIO_FIJO_BSI * cantidad
                return base * 0.26, f"{patologo} → 26% sobre ${base:,.0f} fijo (BSI × {cantidad} + Próstata+{cod})", False
            return subtotal * 0.26, f"{patologo} → 26% (Próstata+{cod})", False
        return subtotal * PCT_GENERAL, f"{patologo} → 13%", False

    return subtotal * PCT_GENERAL, f"{patologo} → 13%", False


def desglosar_iva(imp1, presencia, imp2, condicion, obra_social, pat1="", pat2=""):
    """Desglosa cada importe en exento / gravado 10.5% / gravado 21%.
    Solo calcula IVA para patólogos inscriptos en PATOLOGO_CON_IVA.
    Devuelve dict con 9 columnas de desglose + IVA 10,5% e IVA 21% + Total a facturar."""
    cond = norm(condicion)
    obra = norm(obra_social)

    gravado_21  = (cond == "GRAVADO" and obra == "PARTICULAR")
    gravado_105 = (cond == "GRAVADO" and obra != "PARTICULAR")
    exento      = not (gravado_21 or gravado_105)

    # Separar importes según si el patólogo tiene IVA o no
    imp1_iva  = imp1      if pat1 in PATOLOGO_CON_IVA else 0.0
    imp1_noiv = 0.0       if pat1 in PATOLOGO_CON_IVA else imp1
    pre_iva   = presencia if pat1 in PATOLOGO_CON_IVA else 0.0  # presencias siempre van al pat1
    imp2_iva  = imp2      if pat2 in PATOLOGO_CON_IVA else 0.0
    imp2_noiv = 0.0       if pat2 in PATOLOGO_CON_IVA else imp2

    def asignar(i1, pre, i2):
        if exento:
            return i1, pre, i2, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0
        if gravado_105:
            return 0.0, 0.0, 0.0, i1, pre, i2, 0.0, 0.0, 0.0
        return 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, i1, pre, i2  # gravado_21

    p1_ex, pre_ex, p2_ex, p1_105, pre_105, p2_105, p1_21, pre_21, p2_21 = asignar(imp1_iva, pre_iva, imp2_iva)

    total_105 = p1_105 + pre_105 + p2_105
    total_21  = p1_21  + pre_21  + p2_21
    iva_10    = round(total_105 * 0.105, 2)
    iva_21    = round(total_21  * 0.21,  2)
    total_liq = imp1 + presencia + imp2
    total_fac = round(total_liq + iva_10 + iva_21, 2)

    return {
        "1ra Firma Exento":          round(p1_ex,   2),
        "2da Firma Exento":          round(p2_ex,   2),
        "Presencia Exento":          round(pre_ex,  2),
        "1ra Firma Gravado 10,5%":   round(p1_105,  2),
        "2da Firma Gravado 10,5%":   round(p2_105,  2),
        "Presencia Gravado 10,5%":   round(pre_105, 2),
        "1ra Firma Gravado 21%":     round(p1_21,   2),
        "2da Firma Gravado 21%":     round(p2_21,   2),
        "Presencia Gravado 21%":     round(pre_21,  2),
        "IVA 10,5%":                 iva_10,
        "IVA 21%":                   iva_21,
    }


def procesar(df, monto_alberton=MONTO_FIJO_ALBERTON_DEFAULT):
    alberton_pagados = set()
    resultados = []

    for _, row in df.iterrows():
        subtotal    = pd.to_numeric(row.get("Subtotal", 0), errors="coerce") or 0.0
        cantidad    = max(int(pd.to_numeric(row.get("Cantidad", 1), errors="coerce") or 1), 1)
        cod_fact    = str(row.get("Cod.Fact.", "")).strip()
        organo_raw  = str(row.get("Organo", ""))
        derivante   = row.get("Derivante", "")
        ingreso     = str(row.get("Ingreso", "")).strip()
        condicion   = row.get("Condicion:", "")
        obra_social = row.get("Obra social", "")
        obs         = []

        pat1     = norm(row.get("Patologo", ""))
        pat2_raw = norm(row.get("Segunda Validacion", ""))
        pat2     = "" if pat2_raw in ("NO TIENE", "") else pat2_raw

        cod_es_especial = (
            (cod_fact in CODIGOS_94_PCT and pat1 != "GUSTAVO BARRIENTOS") or
            (pat1 == "GUSTAVO BARRIENTOS" and cod_fact in CODIGOS_BARRIENTOS_FIJO)
        )

        if cod_es_especial and pat2:
            obs.append(f"⚠ Código especial {cod_fact} con 2da firma ({pat2}) → ignorada")
        if not pat1:
            obs.append("⚠ Sin patólogo en primera firma")
        if subtotal == 0:
            obs.append("⚠ Subtotal = 0")

        imp1, regla1, es_pres1 = calcular_importe(pat1, "primera", subtotal, cod_fact, organo_raw, derivante, cantidad)
        imp2, regla2, es_pres2 = calcular_importe(pat2, "segunda", subtotal, cod_fact, organo_raw, derivante, cantidad)
        if cod_es_especial:
            imp2 = 0.0
            es_pres2 = False

        # Separar importe de firma vs presencia
        presencia   = round(imp1, 2) if es_pres1 else 0.0
        imp1_firma  = 0.0            if es_pres1 else round(imp1, 2)
        imp2_firma  = round(imp2, 2) if not es_pres2 else 0.0

        # Valeria Alberton: un único monto fijo por ingreso
        if pat1 == "VALERIA ALBERTON" or pat2 == "VALERIA ALBERTON":
            if ingreso not in alberton_pagados:
                imp1_firma = monto_alberton if pat1 == "VALERIA ALBERTON" else imp1_firma
                imp2_firma = monto_alberton if pat2 == "VALERIA ALBERTON" else imp2_firma
                regla1 = f"Valeria Alberton → ${monto_alberton:,.2f} fijo (ingreso {ingreso})"
                regla2 = regla1
                alberton_pagados.add(ingreso)
                obs.append(f"ℹ Monto fijo Alberton — ingreso {ingreso}")
            else:
                if pat1 == "VALERIA ALBERTON":
                    imp1_firma = 0.0
                    regla1 = f"Valeria Alberton → $0 (ingreso {ingreso} ya liquidado)"
                if pat2 == "VALERIA ALBERTON":
                    imp2_firma = 0.0
                    regla2 = f"Valeria Alberton → $0 (ingreso {ingreso} ya liquidado)"

        total = imp1_firma + imp2_firma + presencia
        regla = f"1ra: {regla1} | 2da: {regla2}" if pat2 and not cod_es_especial else regla1

        # Desglosar IVA por componente
        iva = desglosar_iva(imp1_firma, presencia, imp2_firma, condicion, obra_social, pat1=pat1, pat2=pat2)

        # Preservar todas las columnas originales + agregar liquidación al final
        fila = {col: row.get(col, "") for col in df.columns}
        fila["Patólogo Primera Firma"]     = pat1
        fila["Importe Primera Firma"]      = imp1_firma
        fila["Presencias"]                 = presencia
        fila["Patólogo Segunda Firma"]     = pat2
        fila["Importe Segunda Firma"]      = imp2_firma
        fila["Total Liquidado"]            = round(total, 2)
        fila["1ra Firma Exento"]           = iva["1ra Firma Exento"]
        fila["2da Firma Exento"]           = iva["2da Firma Exento"]
        fila["Presencia Exento"]           = iva["Presencia Exento"]
        fila["1ra Firma Gravado 10,5%"]    = iva["1ra Firma Gravado 10,5%"]
        fila["2da Firma Gravado 10,5%"]    = iva["2da Firma Gravado 10,5%"]
        fila["Presencia Gravado 10,5%"]    = iva["Presencia Gravado 10,5%"]
        fila["1ra Firma Gravado 21%"]      = iva["1ra Firma Gravado 21%"]
        fila["2da Firma Gravado 21%"]      = iva["2da Firma Gravado 21%"]
        fila["Presencia Gravado 21%"]      = iva["Presencia Gravado 21%"]
        fila["IVA 10,5%"]                  = iva["IVA 10,5%"]
        fila["IVA 21%"]                    = iva["IVA 21%"]
        # Total a facturar = Total Liquidado + IVA correspondiente
        fila["Total a facturar"]           = round(total + iva["IVA 10,5%"] + iva["IVA 21%"], 2)
        fila["Regla Aplicada"]             = regla
        fila["Observaciones"]              = " | ".join(obs) if obs else ""
        resultados.append(fila)

    return pd.DataFrame(resultados)


# ─── ESTILOS ─────────────────────────────────────────────────────────────────

AZUL_OSCURO = "1F3864"
AZUL_CLARO  = "D6E4F0"
ROJO_CLARO  = "FCE4D6"
VERDE_CLARO = "E2EFDA"
GRIS        = "F2F2F2"

def hdr(cell, bg=AZUL_OSCURO, fg="FFFFFF"):
    cell.font      = Font(bold=True, color=fg, name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def brd():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def auto_w(ws):
    for col in ws.columns:
        mx = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(mx + 4, 48)

# Columnas originales del Excel + columnas de liquidación al final
COLS_ORIGINALES = [
    "Tipo","Descripcion","Liquidado en","Cod.Fact.","Cod. OS","Cantidad",
    "Ingreso","Fecha","Paciente","Afiliado","DNI","Fec. Nacimiento",
    "Obra social","Condicion:","Derivante","Organo","Matricula",
    "Fecha confirmacion","Precio Unit.","Subtotal",
]
COLS_LIQ = [
    "Patólogo Primera Firma","Importe Primera Firma","Presencias",
    "Patólogo Segunda Firma","Importe Segunda Firma","Total Liquidado",
    "1ra Firma Exento","2da Firma Exento","Presencia Exento",
    "1ra Firma Gravado 10,5%","2da Firma Gravado 10,5%","Presencia Gravado 10,5%",
    "1ra Firma Gravado 21%","2da Firma Gravado 21%","Presencia Gravado 21%",
    "IVA 10,5%","IVA 21%","Total a facturar",
    "Regla Aplicada","Observaciones",
]
COLS_DET = COLS_ORIGINALES + COLS_LIQ
MONEY = {
    "Importe Primera Firma","Presencias","Importe Segunda Firma","Total Liquidado",
    "1ra Firma Exento","2da Firma Exento","Presencia Exento",
    "1ra Firma Gravado 10,5%","2da Firma Gravado 10,5%","Presencia Gravado 10,5%",
    "1ra Firma Gravado 21%","2da Firma Gravado 21%","Presencia Gravado 21%",
    "IVA 10,5%","IVA 21%","Total a facturar","Subtotal","Precio Unit.",
}


def escribir_hoja_detalle(ws, filas_df, titulo_hoja=None, cols_override=None):
    """Escribe filas de detalle en una hoja con formato."""
    base = cols_override if cols_override is not None else COLS_DET
    cols = [c for c in base if c in filas_df.columns]
    ws.append(cols)
    for c in ws[1]: hdr(c)
    ws.row_dimensions[1].height = 30

    for i, row in filas_df.iterrows():
        ws.append([row[c] for c in cols])
        er  = ws.max_row
        obs = bool(row["Observaciones"])
        for j, cell in enumerate(ws[er]):
            cell.border    = brd()
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="center")
            if obs:           cell.fill = PatternFill("solid", start_color=ROJO_CLARO)
            elif er % 2 == 0: cell.fill = PatternFill("solid", start_color=GRIS)
            if cols[j] in MONEY: cell.number_format = '$#,##0.00'

    # Fila de total al pie
    # Fila de total al pie
    ws.append([])
    total_1ra = filas_df["Importe Primera Firma"].sum() if "Importe Primera Firma" in filas_df.columns else 0
    total_2da = filas_df["Importe Segunda Firma"].sum() if "Importe Segunda Firma" in filas_df.columns else 0
    total_gen = filas_df["Total Liquidado"].sum() if "Total Liquidado" in filas_df.columns else 0
    fila_tot = [""] * len(cols)
    for idx, c in enumerate(cols):
        if c == "Paciente": fila_tot[idx] = "TOTAL"
        elif c == "Importe Primera Firma":  fila_tot[idx] = round(total_1ra, 2)
        elif c == "Importe Segunda Firma":  fila_tot[idx] = round(total_2da, 2)
        elif c == "Total Liquidado":        fila_tot[idx] = round(total_gen, 2)
    ws.append(fila_tot)
    er = ws.max_row
    for j, cell in enumerate(ws[er]):
        cell.border = brd()
        cell.font   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        cell.fill   = PatternFill("solid", start_color=AZUL_OSCURO)
        cell.alignment = Alignment(vertical="center", horizontal="center")
        if cols[j] in MONEY: cell.number_format = '$#,##0.00'

    ws.freeze_panes = "A2"
    auto_w(ws)


def construir_excel_general(df_result):
    """Excel completo con todas las hojas."""
    wb  = Workbook()
    wb.remove(wb.active)

    # DETALLE
    ws = wb.create_sheet("Detalle")
    escribir_hoja_detalle(ws, df_result)

    # RESUMEN POR PATÓLOGO
    # ── DATOS BASE para los 3 resúmenes ──────────────────────────────────────────
    cols_desglose = [
        "1ra Firma Exento","Presencia Exento","2da Firma Exento",
        "1ra Firma Gravado 10,5%","Presencia Gravado 10,5%","2da Firma Gravado 10,5%",
        "1ra Firma Gravado 21%","Presencia Gravado 21%","2da Firma Gravado 21%",
        "IVA 10,5%","IVA 21%","Total a facturar",
    ]
    agg_d1 = {"e1":("Importe Primera Firma","count"),"t1":("Importe Primera Firma","sum"),"tp":("Presencias","sum")}
    for c in cols_desglose: agg_d1[c] = (c,"sum")
    g1_all = df_result[df_result["Patólogo Primera Firma"]!=""].groupby("Patólogo Primera Firma").agg(**agg_d1).reset_index().rename(columns={"Patólogo Primera Firma":"P"})
    g2_all = df_result[df_result["Patólogo Segunda Firma"]!=""].groupby("Patólogo Segunda Firma").agg(
        e2=("Importe Segunda Firma","count"), t2=("Importe Segunda Firma","sum")
    ).reset_index().rename(columns={"Patólogo Segunda Firma":"P"})
    res_all = pd.merge(g1_all, g2_all, on="P", how="outer").fillna(0)
    res_all["tot"] = res_all["t1"] + res_all["tp"] + res_all["t2"]
    res_all = res_all.sort_values("tot", ascending=False)

    res_iva   = res_all[res_all["P"].isin(PATOLOGO_CON_IVA)].copy()
    res_noiva = res_all[~res_all["P"].isin(PATOLOGO_CON_IVA)].copy()

    # ── Función para escribir hoja CON IVA (16 columnas) ──────────────────────
    def escribir_resumen_iva(ws_dest, res, titulo_hoja):
        hdrs = [
            "Patólogo",
            "Estudios 1ra Firma","Total exento 1ra Firma ($)","Total gravado 10,5% 1ra Firma ($)","Total gravado 21% 1ra Firma ($)",
            "Presencias","Total exento Presencias ($)","Total gravado 10,5% Presencias ($)","Total gravado 21% Presencias ($)",
            "Estudios 2da Firma","Total exento 2da Firma ($)","Total gravado 10,5% 2da Firma ($)","Total gravado 21% 2da Firma ($)",
            "IVA 10,5% ($)","IVA 21% ($)","Total a facturar ($)",
        ]
        ws_dest.append(hdrs)
        COLOR_1RA="1F3864"; COLOR_PRES="375623"; COLOR_2DA="843C0C"; COLOR_IVA="7B2C9E"
        grupos={0:COLOR_1RA,1:COLOR_1RA,2:COLOR_1RA,3:COLOR_1RA,4:COLOR_1RA,
                5:COLOR_PRES,6:COLOR_PRES,7:COLOR_PRES,8:COLOR_PRES,
                9:COLOR_2DA,10:COLOR_2DA,11:COLOR_2DA,12:COLOR_2DA,
                13:COLOR_IVA,14:COLOR_IVA,15:COLOR_IVA}
        for j,cell in enumerate(ws_dest[1]):
            cell.font=Font(bold=True,color="FFFFFF",name="Arial",size=9)
            cell.fill=PatternFill("solid",start_color=grupos.get(j,COLOR_1RA))
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
            cell.border=brd()
        ws_dest.row_dimensions[1].height=50
        MONEY_R={2,3,4,6,7,8,10,11,12,13,14,15}
        for _,r in res.iterrows():
            ws_dest.append([
                r["P"],
                int(r["e1"]),round(r["1ra Firma Exento"],2),round(r["1ra Firma Gravado 10,5%"],2),round(r["1ra Firma Gravado 21%"],2),
                round(r["tp"],2),round(r["Presencia Exento"],2),round(r["Presencia Gravado 10,5%"],2),round(r["Presencia Gravado 21%"],2),
                int(r["e2"]),round(r["2da Firma Exento"],2),round(r["2da Firma Gravado 10,5%"],2),round(r["2da Firma Gravado 21%"],2),
                round(r["IVA 10,5%"],2),round(r["IVA 21%"],2),round(r["Total a facturar"],2),
            ])
            er=ws_dest.max_row
            for j,cell in enumerate(ws_dest[er]):
                cell.border=brd(); cell.font=Font(name="Arial",size=9); cell.alignment=Alignment(vertical="center")
                if er%2==0: cell.fill=PatternFill("solid",start_color=AZUL_CLARO)
                if j in MONEY_R: cell.number_format="$#,##0.00"
        ft=ws_dest.max_row+1
        tots=["TOTAL GENERAL",f"=SUM(B2:B{ft-1})"]
        for ci in range(3,17): tots.append(f"=SUM({get_column_letter(ci)}2:{get_column_letter(ci)}{ft-1})")
        ws_dest.append(tots)
        for j,cell in enumerate(ws_dest[ft]):
            hdr(cell); cell.border=brd()
            if j in MONEY_R: cell.number_format="$#,##0.00"
        ws_dest.freeze_panes="A2"; auto_w(ws_dest)

    # ── Función para escribir hoja SIN IVA (7 columnas simples) ───────────────
    def escribir_resumen_noiva(ws_dest, res):
        hdrs=["Patólogo","Estudios 1ra Firma","Total 1ra Firma ($)","Presencias ($)","Estudios 2da Firma","Total 2da Firma ($)","Total General ($)"]
        ws_dest.append(hdrs)
        for c in ws_dest[1]: hdr(c)
        ws_dest.row_dimensions[1].height=30
        for _,r in res.iterrows():
            ws_dest.append([r["P"],int(r["e1"]),round(r["t1"],2),round(r["tp"],2),int(r["e2"]),round(r["t2"],2),round(r["tot"],2)])
            er=ws_dest.max_row
            for j,cell in enumerate(ws_dest[er]):
                cell.border=brd(); cell.font=Font(name="Arial",size=9); cell.alignment=Alignment(vertical="center")
                if er%2==0: cell.fill=PatternFill("solid",start_color=AZUL_CLARO)
                if j in (2,3,5,6): cell.number_format="$#,##0.00"
        ft=ws_dest.max_row+1
        ws_dest.append(["TOTAL GENERAL",f"=SUM(B2:B{ft-1})",f"=SUM(C2:C{ft-1})",f"=SUM(D2:D{ft-1})",
                        f"=SUM(E2:E{ft-1})",f"=SUM(F2:F{ft-1})",f"=SUM(G2:G{ft-1})"])
        for j,cell in enumerate(ws_dest[ft]):
            hdr(cell); cell.border=brd()
            if j in (2,3,5,6): cell.number_format="$#,##0.00"
        ws_dest.freeze_panes="A2"; auto_w(ws_dest)

    # ── HOJA 1: Resumen con IVA ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Resumen con IVA")
    escribir_resumen_iva(ws2, res_iva, "Resumen con IVA")

    # ── HOJA 2: Resumen sin IVA ───────────────────────────────────────────────
    ws2b = wb.create_sheet("Resumen sin IVA")
    escribir_resumen_noiva(ws2b, res_noiva)

    # ── HOJA 3: Resumen General (todos) ──────────────────────────────────────
    ws2c = wb.create_sheet("Resumen General")
    escribir_resumen_iva(ws2c, res_all, "Resumen General")

    # INCONSISTENCIAS
    ws3    = wb.create_sheet("Inconsistencias")
    df_inc = df_result[df_result["Observaciones"] != ""].copy()
    c3     = ["Paciente","Fecha","Ingreso","Cod.Fact.","Patólogo Primera Firma","Patólogo Segunda Firma","Observaciones"]
    ws3.append(c3)
    for c in ws3[1]: hdr(c, bg="C00000")
    ws3.row_dimensions[1].height = 30
    if df_inc.empty:
        ws3.append(["✅ Sin inconsistencias detectadas"] + [""] * 6)
    else:
        for _, row in df_inc.iterrows():
            ws3.append([row[c] for c in c3])
            er = ws3.max_row
            for cell in ws3[er]:
                cell.border    = brd()
                cell.font      = Font(name="Arial", size=9, color="7B0000")
                cell.fill      = PatternFill("solid", start_color=ROJO_CLARO)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws3.freeze_panes = "A2"; auto_w(ws3)

    # TOTALES GENERALES
    ws4 = wb.create_sheet("Totales Generales")
    ws4.column_dimensions["A"].width = 42
    ws4.column_dimensions["B"].width = 22

    def titulo(t):
        ws4.append([t, ""])
        ws4.merge_cells(f"A{ws4.max_row}:B{ws4.max_row}")
        hdr(ws4.cell(ws4.max_row, 1))

    def kv(label, value, fmt="$#,##0.00", bold=False):
        ws4.append([label, value]); r = ws4.max_row
        ws4[f"A{r}"].font = Font(name="Arial", size=10, bold=bold)
        ws4[f"B{r}"].font = Font(name="Arial", size=10, bold=bold)
        ws4[f"B{r}"].number_format = fmt
        ws4[f"B{r}"].alignment = Alignment(horizontal="right")
        for col in ["A","B"]: ws4[f"{col}{r}"].border = brd()

    titulo("📊 RESUMEN GENERAL DE LIQUIDACIÓN")
    ws4.append([])
    kv("Total estudios procesados",    len(df_result),                       fmt="0")
    kv("Total subtotal facturado",      df_result["Subtotal"].sum())
    kv("Total liquidado primera firma", df_result["Importe Primera Firma"].sum())
    kv("Total presencias",               df_result["Presencias"].sum())
    kv("Total liquidado segunda firma", df_result["Importe Segunda Firma"].sum())
    kv("TOTAL GENERAL LIQUIDADO",       df_result["Total Liquidado"].sum(),  bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf, len(df_inc)


def construir_excel_patologo(nombre_patologo, filas_1ra, filas_2da):
    """Excel individual para un patólogo — solo muestra sus propios importes, el resto en 0."""
    wb = Workbook()
    wb.remove(wb.active)

    total_general = 0.0

    COLS_SIN_OBS = [c for c in COLS_DET if c not in ("Regla Aplicada", "Observaciones")]

    # Hoja 1ra firma: solo su importe, sin regla ni observaciones
    # Para Valeria Alberton: eliminar filas duplicadas del mismo ingreso (importe = 0)
    if not filas_1ra.empty:
        df_1ra = filas_1ra.copy()
        df_1ra["Importe Segunda Firma"] = 0.0
        df_1ra["Total Liquidado"]       = df_1ra["Importe Primera Firma"] + df_1ra["Presencias"]
        if nombre_patologo == "VALERIA ALBERTON":
            df_1ra = df_1ra[(df_1ra["Importe Primera Firma"] > 0) | (df_1ra["Presencias"] > 0)]
        ws1 = wb.create_sheet("Primera Firma")
        escribir_hoja_detalle(ws1, df_1ra, cols_override=COLS_SIN_OBS)
        total_general += df_1ra["Importe Primera Firma"].sum()

    # Hoja 2da firma: solo su importe, sin regla ni observaciones
    if not filas_2da.empty:
        df_2da = filas_2da.copy()
        df_2da["Importe Primera Firma"] = 0.0
        df_2da["Presencias"]            = 0.0
        df_2da["Total Liquidado"]       = df_2da["Importe Segunda Firma"]
        ws2 = wb.create_sheet("Segunda Firma")
        escribir_hoja_detalle(ws2, df_2da, cols_override=COLS_SIN_OBS)
        total_general += df_2da["Importe Segunda Firma"].sum()

    # Hoja resumen del patólogo
    wsr = wb.create_sheet("Resumen")
    wsr.column_dimensions["A"].width = 40
    wsr.column_dimensions["B"].width = 22

    def titulo(t, bg=AZUL_OSCURO):
        wsr.append([t, ""])
        wsr.merge_cells(f"A{wsr.max_row}:B{wsr.max_row}")
        hdr(wsr.cell(wsr.max_row, 1), bg=bg)

    def kv(label, value, fmt="$#,##0.00", bold=False):
        wsr.append([label, value]); r = wsr.max_row
        wsr[f"A{r}"].font = Font(name="Arial", size=10, bold=bold)
        wsr[f"B{r}"].font = Font(name="Arial", size=10, bold=bold)
        wsr[f"B{r}"].number_format = fmt
        wsr[f"B{r}"].alignment = Alignment(horizontal="right")
        for col in ["A","B"]: wsr[f"{col}{r}"].border = brd()

    titulo(f"🔬 LIQUIDACIÓN — {nombre_patologo}")
    wsr.append([])

    if not filas_1ra.empty:
        kv("Estudios como 1ra firma",            len(filas_1ra),                                    fmt="0")
        kv("Total exento 1ra firma",              filas_1ra["1ra Firma Exento"].sum())
        kv("Total gravado 10,5% 1ra firma",       filas_1ra["1ra Firma Gravado 10,5%"].sum())
        kv("Total gravado 21% 1ra firma",         filas_1ra["1ra Firma Gravado 21%"].sum())
        kv("Presencias exento",                   filas_1ra["Presencia Exento"].sum())
        kv("Presencias gravado 10,5%",            filas_1ra["Presencia Gravado 10,5%"].sum())
        kv("Presencias gravado 21%",              filas_1ra["Presencia Gravado 21%"].sum())
        wsr.append([])

    if not filas_2da.empty:
        kv("Estudios como 2da firma",            len(filas_2da),                                    fmt="0")
        kv("Total exento 2da firma",              filas_2da["2da Firma Exento"].sum())
        kv("Total gravado 10,5% 2da firma",       filas_2da["2da Firma Gravado 10,5%"].sum())
        kv("Total gravado 21% 2da firma",         filas_2da["2da Firma Gravado 21%"].sum())
        wsr.append([])

    iva_10_total = (filas_1ra["IVA 10,5%"].sum() if not filas_1ra.empty else 0) +                   (filas_2da["IVA 10,5%"].sum() if not filas_2da.empty else 0)
    iva_21_total = (filas_1ra["IVA 21%"].sum() if not filas_1ra.empty else 0) +                   (filas_2da["IVA 21%"].sum() if not filas_2da.empty else 0)
    taf_total    = (filas_1ra["Total a facturar"].sum() if not filas_1ra.empty else 0) +                   (filas_2da["Total a facturar"].sum() if not filas_2da.empty else 0)

    kv("TOTAL A COBRAR",    total_general, bold=True)
    if iva_10_total > 0 or iva_21_total > 0:
        wsr.append([])
        kv("IVA 10,5%",         iva_10_total)
        kv("IVA 21%",           iva_21_total)
        kv("TOTAL A FACTURAR",  taf_total,    bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def construir_zip_patologo(df_result, nombre_base):
    """Genera un ZIP con un Excel por cada patólogo."""
    zip_buf = io.BytesIO()
    patologos = set(df_result["Patólogo Primera Firma"].unique()) | set(df_result["Patólogo Segunda Firma"].unique())
    patologos.discard("")

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for pat in sorted(patologos):
            filas_1ra = df_result[df_result["Patólogo Primera Firma"] == pat].copy()
            filas_2da = df_result[df_result["Patólogo Segunda Firma"] == pat].copy()

            # Solo incluir si tiene algo liquidado
            total = filas_1ra["Importe Primera Firma"].sum() + filas_2da["Importe Segunda Firma"].sum()

            excel_pat = construir_excel_patologo(pat, filas_1ra, filas_2da)
            nombre_archivo = f"Optimi_{pat.replace(' ', '_').replace('/', '-')}_{nombre_base}.xlsx"
            zf.writestr(nombre_archivo, excel_pat.read())

    zip_buf.seek(0)
    return zip_buf, len(patologos)


# ─── STREAMLIT UI ─────────────────────────────────────────────────────────────

st.set_page_config(page_title="Liquidación Patólogos", page_icon="🔬", layout="wide")

st.markdown("""
<style>
.block-container { padding-top: 2rem; }
.stButton>button {
    background-color: #1F3864; color: white;
    border-radius: 8px; padding: 0.5rem 2rem;
    font-size: 1rem; font-weight: bold; width: 100%;
}
.stButton>button:hover { background-color: #2e4f8a; }
.stDownloadButton>button {
    background-color: #1a7a4a; color: white;
    border-radius: 8px; padding: 0.6rem 2rem;
    font-size: 1rem; font-weight: bold; width: 100%;
}
.stDownloadButton>button:hover { background-color: #145e38; }
.metric-card {
    background: white; border-radius: 12px;
    padding: 1.2rem 1rem; text-align: center;
    box-shadow: 0 2px 8px rgba(0,0,0,0.08);
    border-top: 4px solid #1F3864;
}
.metric-value { font-size: 1.5rem; font-weight: 800; color: #1F3864; }
.metric-label { font-size: 0.8rem; color: #888; margin-top: 4px; text-transform: uppercase; letter-spacing: 0.05em; }
.alerta { background:#fff3cd; border-left:4px solid #ffc107; padding:0.8rem 1rem; border-radius:6px; margin:0.5rem 0; }
.descarga-box { background:#f0f7f0; border:2px solid #1a7a4a; border-radius:10px; padding:1.2rem; margin:0.5rem 0; }
</style>
""", unsafe_allow_html=True)

col_logo, col_titulo = st.columns([1, 8])
with col_logo:
    st.markdown("<div style='font-size:3rem;text-align:center'>🔬</div>", unsafe_allow_html=True)
with col_titulo:
    st.markdown("## Liquidación de Patólogos")
    st.markdown("<span style='color:#666;font-size:0.95rem'>Subí el archivo Excel consolidado y descargá la liquidación calculada automáticamente</span>", unsafe_allow_html=True)

st.divider()

uploaded = st.file_uploader(
    "📂 Seleccioná el archivo Excel",
    type=["xlsx", "xls"],
    help="El archivo debe tener una hoja llamada LIQUIDACION o BASE"
)

if uploaded:
    try:
        xf = pd.ExcelFile(uploaded)
        hojas = xf.sheet_names
        hoja_datos = None
        for candidata in ["LIQUIDACION", "BASE"]:
            if candidata in hojas:
                hoja_datos = candidata
                break

        if hoja_datos is None:
            st.error(f"❌ No se encontró una hoja válida. Hojas disponibles: {', '.join(hojas)}")
            st.stop()

        with st.spinner("Leyendo archivo..."):
            df = pd.read_excel(uploaded, sheet_name=hoja_datos, header=0, dtype=str)
            df.columns = [str(c).replace("\xa0", " ").strip() for c in df.columns]
            df["Subtotal"] = pd.to_numeric(df["Subtotal"], errors="coerce").fillna(0)
            # Eliminar filas con Cod. OS = 0, vacío o nulo
            filas_antes = len(df)
            df = df[~df["Cod. OS"].fillna("").str.strip().isin(["0", "0.0", ""])]
            df = df.reset_index(drop=True)
            filas_eliminadas = filas_antes - len(df)

        nombre_base = uploaded.name.replace(".xlsx","").replace(".xls","")
        msg = f"✅ Archivo cargado — hoja **{hoja_datos}** — {len(df):,} estudios"
        if filas_eliminadas > 0:
            msg += f" ({filas_eliminadas} filas con Cod. OS=0 eliminadas)"
        st.success(msg)

    except Exception as e:
        st.error(f"❌ Error al leer el archivo: {e}")
        st.stop()

    st.divider()
    st.subheader("⚙️ Parámetros del mes")
    col_alb, col_vacio = st.columns([1, 2])
    with col_alb:
        monto_alberton = st.number_input(
            "Monto Valeria Alberton ($)",
            min_value=0.0,
            value=MONTO_FIJO_ALBERTON_DEFAULT,
            step=100.0,
            format="%.2f",
            help=f"Monto fijo por ingreso para este mes. Se actualiza según el aumento OSDE acumulado. Valor base: ${MONTO_FIJO_ALBERTON_DEFAULT:,.2f}"
        )
    st.divider()

    if st.button("⚡ Calcular liquidación"):
        prog = st.progress(0, text="Procesando estudios...")
        with st.spinner(""):
            df_result = procesar(df, monto_alberton=monto_alberton)
            prog.progress(50, text="Generando Excel general...")
            excel_buf, n_inc = construir_excel_general(df_result)
            prog.progress(80, text="Generando archivos por patólogo...")
            zip_buf, n_patologos = construir_zip_patologo(df_result, nombre_base)
            prog.progress(100, text="¡Listo!")

        st.divider()
        st.subheader("📊 Resumen")

        total_liq = df_result["Total Liquidado"].sum()
        total_sub = df_result["Subtotal"].sum()
        total_1ra = df_result["Importe Primera Firma"].sum()
        total_2da = df_result["Importe Segunda Firma"].sum()

        c1, c2, c3, c4, c5 = st.columns(5)
        for col, label, value in [
            (c1, "Estudios",           f"{len(df_result):,}"),
            (c2, "Subtotal facturado",  f"${total_sub:,.0f}"),
            (c3, "Total 1ra firma",    f"${total_1ra:,.0f}"),
            (c4, "Total 2da firma",    f"${total_2da:,.0f}"),
            (c5, "TOTAL LIQUIDADO",    f"${total_liq:,.0f}"),
        ]:
            col.markdown(f"""
                <div class='metric-card'>
                    <div class='metric-value'>{value}</div>
                    <div class='metric-label'>{label}</div>
                </div>
            """, unsafe_allow_html=True)

        st.markdown("")

        if n_inc > 0:
            subtotal_cero = df_result[df_result["Observaciones"].str.contains("Subtotal = 0", na=False)]
            cod_esp_2da   = df_result[df_result["Observaciones"].str.contains("Código especial", na=False)]
            sin_pat       = df_result[df_result["Observaciones"].str.contains("Sin patólogo", na=False)]
            st.markdown(f"<div class='alerta'>⚠️ Se detectaron <b>{n_inc} inconsistencias</b> — revisalas en la hoja <b>Inconsistencias</b> del Excel general.</div>", unsafe_allow_html=True)
            cols_inc = st.columns(3)
            if len(subtotal_cero): cols_inc[0].metric("Subtotal = 0", len(subtotal_cero))
            if len(cod_esp_2da):   cols_inc[1].metric("Códigos especiales con 2da firma", len(cod_esp_2da))
            if len(sin_pat):       cols_inc[2].metric("Sin patólogo", len(sin_pat))

        st.divider()
        st.subheader("👩‍⚕️ Totales por patólogo")

        g1_st = df_result[df_result["Patólogo Primera Firma"] != ""].groupby("Patólogo Primera Firma").agg(
            pf=("Importe Primera Firma","sum"), pres=("Presencias","sum"),
            iv10=("IVA 10,5%","sum"), iv21=("IVA 21%","sum"), taf=("Total a facturar","sum")
        ).reset_index().rename(columns={"Patólogo Primera Firma":"Patólogo"})
        g2_st = df_result[df_result["Patólogo Segunda Firma"] != ""].groupby("Patólogo Segunda Firma").agg(
            sf=("Importe Segunda Firma","sum")
        ).reset_index().rename(columns={"Patólogo Segunda Firma":"Patólogo"})
        resumen = pd.merge(g1_st, g2_st, on="Patólogo", how="outer").fillna(0)
        resumen["Total ($)"] = resumen["pf"] + resumen["pres"] + resumen["sf"]
        resumen = resumen.sort_values("Total ($)", ascending=False).reset_index(drop=True)
        resumen = resumen.rename(columns={
            "pf":"1ra Firma ($)", "pres":"Presencias ($)", "sf":"2da Firma ($)",
            "iv10":"IVA 10,5% ($)", "iv21":"IVA 21% ($)", "taf":"Total a facturar ($)"
        })
        resumen = resumen[["Patólogo","1ra Firma ($)","Presencias ($)","2da Firma ($)","Total ($)","IVA 10,5% ($)","IVA 21% ($)","Total a facturar ($)"]]
        fila_total = pd.DataFrame([{c: resumen[c].sum() if c != "Patólogo" else "TOTAL GENERAL" for c in resumen.columns}])
        resumen_display = pd.concat([resumen, fila_total], ignore_index=True)
        for col in resumen.columns[1:]:
            resumen_display[col] = resumen_display[col].apply(lambda x: f"${float(x):,.2f}" if x != "" else "")
        st.dataframe(resumen_display, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("⬇️ Descargas")

        col_d1, col_d2 = st.columns(2)

        with col_d1:
            st.markdown("<div class='descarga-box'>", unsafe_allow_html=True)
            st.markdown("**📋 Excel general completo**")
            st.markdown("Incluye todas las hojas: Detalle, Resumen, Inconsistencias y Totales.")
            st.download_button(
                label="⬇️ Descargar Excel general",
                data=excel_buf,
                file_name=f"liquidacion_{nombre_base}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            st.markdown("</div>", unsafe_allow_html=True)

        with col_d2:
            st.markdown("<div class='descarga-box'>", unsafe_allow_html=True)
            st.markdown(f"**🗂 ZIP con {n_patologos} archivos individuales**")
            st.markdown("Un Excel por patólogo con sus estudios de 1ra y 2da firma + resumen.")
            st.download_button(
                label="⬇️ Descargar ZIP por patólogo",
                data=zip_buf,
                file_name=f"liquidacion_por_patologo_{nombre_base}.zip",
                mime="application/zip",
                use_container_width=True,
            )
            st.markdown("</div>", unsafe_allow_html=True)

else:
    st.info("👆 Subí el archivo Excel para comenzar.")
    with st.expander("ℹ️ ¿Cómo funciona?"):
        st.markdown("""
        1. **Subís** el archivo Excel consolidado de patólogos
        2. **Hacés clic** en *Calcular liquidación*
        3. La app aplica todas las reglas automáticamente
        4. **Descargás** dos archivos:
           - **Excel general**: Detalle, Resumen, Inconsistencias y Totales
           - **ZIP individual**: un Excel por patólogo con sus estudios y total a cobrar
        """)

st.divider()
st.caption("Liquidación de Patólogos · Reglas 2025 · Desarrollado con Streamlit")
